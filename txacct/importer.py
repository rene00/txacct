from txacct.model import (
    StatisticalArea3,
    StatisticalArea4,
    State,
    Postcode,
    OrganisationSource,
    Organisation,
    ANZSIC,
    BusinessCode,
)
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook
from pathlib import Path


class Importer:
    def __validate(self):
        """Validate init kwargs"""
        if self.organisation_data:
            if not self.filename:
                raise Exception("filename must be set if importing organisation data")
            if not self.worksheet:
                raise Exception("worksheet must be set if importing organisation data")
            if not self.organisation_source:
                raise Exception(
                    "organisation_source must be set if importing organisation data"
                )

        if self.postcode_data is None or self.organisation_data is None:
            raise Exception("postcode_data must be set")

    def __init__(self, session, **kwargs) -> None:
        self.session = session
        self.postcode_data: dict = kwargs.get("postcode_data", {})

        self.organisation_data: bool = kwargs.get("organisation_data", False)
        self.filename: Path = kwargs.get("filename", None)
        self.worksheet: str = kwargs.get("worksheet", None)
        self.organisation_source: str = kwargs.get("organisation_source", None)
        self.print_headers: bool = kwargs.get("print_headers", False)

        self.__validate()

    def __import_postcode_data(self):
        for i in self.postcode_data:
            sa3 = None
            if i.get("sa3") != "" and i.get("sa3name") != "":
                sa3 = StatisticalArea3(
                    code=i.get("sa3"),
                    name=i.get("sa3name"),
                )
                self.session.add(sa3)
                try:
                    self.session.commit()
                except IntegrityError:
                    self.session.rollback()
                    sa3 = (
                        self.session.query(StatisticalArea3)
                        .filter_by(code=int(sa3.code), name=i.get("sa3name"))
                        .one()
                    )

            sa4 = None
            if i.get("sa4") != "" and i.get("sa4name") != "":
                sa4 = StatisticalArea4(
                    code=i.get("sa4"),
                    name=i.get("sa4name"),
                )
                self.session.add(sa4)
                try:
                    self.session.commit()
                except IntegrityError:
                    self.session.rollback()
                    sa4 = (
                        self.session.query(StatisticalArea4)
                        .filter_by(code=int(sa4.code), name=sa4.name)
                        .one()
                    )

            state = None
            if i.get("state") != "":
                state = State(name=i.get("state"))
                self.session.add(state)
                try:
                    self.session.commit()
                except IntegrityError:
                    self.session.rollback()
                    state = self.session.query(State).filter_by(name=state.name).one()

            postcode: Postcode = Postcode(
                postcode=i.get("postcode"),
                locality=i.get("locality"),
                state=state,
                sa3=sa3,
                sa4=sa4,
            )

            self.session.add(postcode)
            try:
                self.session.commit()
            except IntegrityError:
                self.session.rollback()

    def __business_code(self, company: dict) -> BusinessCode:
        business_code = BusinessCode(
            code=company["BUSCODE"], description=company["BUSINESS_DESCRIPTION"]
        )
        self.session.add(business_code)
        try:
            self.session.commit()
        except IntegrityError:
            self.session.rollback()
            row = self.session.execute(
                select(BusinessCode).where(
                    BusinessCode.code == company["BUSCODE"],
                    BusinessCode.description == company["BUSINESS_DESCRIPTION"],
                )
            ).one()
            business_code = row[0]
        finally:
            return business_code

    def __anzsic(self, company: dict) -> ANZSIC | None:
        """ Import ANZSIC data 

        This is disabled for now. Data provider does not include
        ANZSIC-DESCRIPTION so ANZSIC dataset needs to be sourced first.

        if not company.keys() & {"ANZSIC-CODE", "ANZSIC-DESCRIPTION"}:
            return None
        anzsic = ANZSIC(
            code=company["ANZSIC-CODE"],
            description=company["ANZSIC-DESCRIPTION"],
        )
        self.session.add(anzsic)
        try:
            self.session.commit()
        except IntegrityError:
            self.session.rollback()
            row = self.session.execute(
                select(ANZSIC).where(
                    ANZSIC.code == company["ANZSIC-CODE"],
                    ANZSIC.description == company["ANZSIC-DESCRIPTION"],
                )
            ).one()
            anzsic = row[0]
        finally:
            return anzsic
        """

        return None


    def __postcode(self, company: dict) -> Postcode | None:
        if not company.keys() & {"POSTCODE", "LOCATION"}:
            return None
        return self.session.scalars(
            select(Postcode).where(
                Postcode.postcode == company["POSTCODE"],
                Postcode.locality == company["LOCATION"],
            )
        ).first()

    def __organisation_source(self) -> OrganisationSource:
        org_source = OrganisationSource(name=self.organisation_source)
        self.session.add(org_source)
        try:
            self.session.commit()
        except IntegrityError:
            self.session.rollback()
            row = self.session.execute(
                select(OrganisationSource).where(
                    OrganisationSource.name == self.organisation_source
                )
            ).one()
            org_source = row[0]
        finally:
            return org_source

    def __organisation_data_headers(self) -> dict:
        """ Headers that are found within the Data Provider worksheet

        Items that are hashed out are not imported. """
        return {
            1:"ID_ORGANISATION",
            # 2:"RECORD_DEFUNCT_RISK",
            3:"BUSINESS_DESCRIPTION",
            4:"BUSCODE",
            5:"ORGANISATION",
            6:"ADDRESS",
            7:"LOCATION",
            8:"POSTCODE",
            #9:"STATE",
            #10:"REGION",
            #11:"PHONE",
            #12:"MOBILE",
            #13:"FREECALL",
            #14:"FAX",
            #15:"EMAIL",
            #16:"EMAIL-2",
            #17:"WEBSITE",
            #18:"FACEBOOK",
            #19:"TWITTER",
            #20:"LINKEDIN",
            #21:"EMPLOYEES",
            #22:"REVENUE-$M",
            #23:"YEAR-ESTABLISHED",
            #24:"CONTACT-NAME",
            #25:"CONTACT-FIRST-NAME",
            #26:"CONTACT-JOB-TITLE",
            27:"ABN",
            #28:"ABN-STATUS",
            #29:"STATUS-DATE",
            #30:"ENTITY-TYPE-CODE",
            #31:"ANZSIC-CODE",
            #32:"LATITUDE",
            #33:"LONGITUDE",
            #34:"MAPLINK",
            #35:"ID-ORG",
        }

    def __import_organisation_data(self):
        wb = load_workbook(filename=self.filename)
        ws = wb[self.worksheet]

        headers = self.__organisation_data_headers()

        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                if not self.print_headers:
                    continue
                return

            organisation = Organisation(
                organisation_source=self.__organisation_source()
            )

            buscode = {}
            anzsic = {}
            postcode = {}

            for cell in row:
                if cell.column not in headers:
                    continue
                header = headers[cell.column]
                match header:
                    case "ID_ORGANISATION":
                        organisation.source_id = cell.value
                    case "ORGANISATION":
                        organisation.name = cell.value
                    case "ABN":
                        organisation.abn = cell.value
                    case "ADDRESS":
                        organisation.address = cell.value
                    case "BUSCODE":
                        buscode["BUSCODE"] = cell.value
                    case "BUSINESS_DESCRIPTION":
                        buscode["BUSINESS_DESCRIPTION"] = cell.value
                    case "ANZSIC-CODE":
                        anzsic["ANZSIC-CODE"] = cell.value
                    case "ANZSIC-DESCRIPTION":
                        anzsic["ANZSIC-DESCRIPTION"] = cell.value
                    case "POSTCODE":
                        postcode["POSTCODE"] = cell.value
                    case "LOCATION":
                        postcode["LOCATION"] = cell.value

            organisation.business_code=self.__business_code(buscode)
            organisation.anzsic=self.__anzsic(anzsic)
            organisation.postcode=self.__postcode(postcode)
            self.session.add(organisation)
            try:
                self.session.commit()
            except IntegrityError:
                self.session.rollback()
        return

    def do(self) -> None:
        """import locality data into db"""

        if self.postcode_data:
            self.__import_postcode_data()

        if self.organisation_data:
            self.__import_organisation_data()
